"""
Generate PinTable.xlsx for CommBench (STM32G071RBT6 LQFP64).
Phase 1 active pins + Phase 2+ RESERVED notes + LQFP64 full pinout.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# STM32G071RBT6 LQFP64 pinout (DS12232 Rev 6, Figure 17)
LQFP64_PINS = [
    (1, "VBAT"),
    (2, "PC13"),
    (3, "PC14-OSC32_IN"),
    (4, "PC15-OSC32_OUT"),
    (5, "PF0-OSC_IN"),
    (6, "PF1-OSC_OUT"),
    (7, "NRST"),
    (8, "VSSA/VREF-"),
    (9, "VDDA/VREF+"),
    (10, "PA0"),
    (11, "PA1"),
    (12, "PA2"),
    (13, "PA3"),
    (14, "PA4"),
    (15, "PA5"),
    (16, "PA6"),
    (17, "PA7"),
    (18, "PC4"),
    (19, "PC5"),
    (20, "PB0"),
    (21, "PB1"),
    (22, "PB2"),
    (23, "PB10"),
    (24, "PB11"),
    (25, "VSS"),
    (26, "VDD"),
    (27, "PB12"),
    (28, "PB13"),
    (29, "PB14"),
    (30, "PB15"),
    (31, "PC6"),
    (32, "PC7"),
    (33, "PC8"),
    (34, "PC9"),
    (35, "PA8"),
    (36, "PA9"),
    (37, "PA10"),
    (38, "PA11 [PA9]"),
    (39, "PA12 [PA10]"),
    (40, "PA13 (SWDIO)"),
    (41, "PA14 (SWCLK)"),
    (42, "PA15"),
    (43, "PC10"),
    (44, "PC11"),
    (45, "PC12"),
    (46, "PD0"),
    (47, "PD1"),
    (48, "PD2"),
    (49, "PD3"),
    (50, "PD4"),
    (51, "PD5"),
    (52, "PD6"),
    (53, "PB3"),
    (54, "PB4"),
    (55, "PB5"),
    (56, "PB6"),
    (57, "PB7"),
    (58, "PB8"),
    (59, "PB9"),
    (60, "VSS"),
    (61, "VDD"),
    (62, "PD8"),
    (63, "PD9"),
    (64, "BOOT0"),
]


def sanity_check():
    """Verify pinout matches DS12232 Rev 6 Figure 17 LQFP64."""
    assert len(LQFP64_PINS) == 64, f"Expected 64 pins, got {len(LQFP64_PINS)}"
    pins = {p[0]: p[1] for p in LQFP64_PINS}
    assert sorted(pins.keys()) == list(range(1, 65)), "Pin numbers not contiguous 1..64"
    expected = {
        1: "VBAT",
        2: "PC13",
        7: "NRST",
        8: "VSSA/VREF-",
        9: "VDDA/VREF+",
        10: "PA0",
        12: "PA2",
        13: "PA3",
        15: "PA5",
        25: "VSS",
        26: "VDD",
        40: "PA13 (SWDIO)",
        41: "PA14 (SWCLK)",
        58: "PB8",
        59: "PB9",
        60: "VSS",
        61: "VDD",
        64: "BOOT0",
    }
    for pin, name in expected.items():
        actual = pins[pin]
        assert actual == name, f"Pin {pin} should be {name!r}, got {actual!r}"
    print("Sanity check passed: pin map matches DS12232 Rev 6 Figure 17")


# Phase 1 — active (灌入 .ioc 的)
PHASE1 = {
    "PA2":  ("LPUART1_TX",  "LPUART1", "VCP_TX",       "ST-Link VCP TX (host SCPI command)"),
    "PA3":  ("LPUART1_RX",  "LPUART1", "VCP_RX",       "ST-Link VCP RX (host SCPI command)"),
    "PA5":  ("GPIO_Output", "GPIO",    "LD4",          "板載狀態 LED (busy / error indicator)"),
    "PB8":  ("I2C1_SCL",    "I2C1",    "I2C1_SCL",     "DUT I2C bus clock (4.7k pull-up to 3V3)"),
    "PB9":  ("I2C1_SDA",    "I2C1",    "I2C1_SDA",     "DUT I2C bus data (4.7k pull-up to 3V3)"),
    "PC13": ("GPXTI13",     "GPIO/EXTI","B1_USER_BTN", "板載 B1 按鈕 (EXTI falling edge, 觸發 I2C scan)"),
}

# Phase 2+ — reserved (在表上註記，不灌入 .ioc)
# 設計原則：SPI 移到 PB3/PB4/PB5 alt pins，跟 ADC 完全分開
PHASE2_RESERVED = {
    # DUT UART (USART1) — 跟 ADC_IN14/IN15 物理共腳，捨棄 ADC 給 UART
    "PC4":  ("USART1_TX (Phase 2)",  "DUT UART TX (棄 ADC_IN14)"),
    "PC5":  ("USART1_RX (Phase 2)",  "DUT UART RX (棄 ADC_IN15)"),
    # DUT SPI1 — 走 PB3/PB4/PB5 alt pins，整段避開 ADC 區
    "PB3":  ("SPI1_SCK (Phase 2)",   "DUT SPI clock"),
    "PB4":  ("SPI1_MISO (Phase 2)",  "DUT SPI MISO"),
    "PB5":  ("SPI1_MOSI (Phase 2)",  "DUT SPI MOSI"),
    "PB12": ("SPI1_CS0 (Phase 2)",   "SPI software CS #0"),
    "PB13": ("SPI1_CS1 (Phase 2)",   "SPI software CS #1"),
    "PB14": ("SPI1_CS2 (Phase 2)",   "SPI software CS #2"),
    "PB15": ("SPI1_CS3 (Phase 2)",   "SPI software CS #3"),
    # ADC1 — 6 channel 完全乾淨區（PA4 已讓給 DAC，PC1 補位）
    "PA0":  ("ADC1_IN0 (Phase 2)",   "ADC channel 0"),
    "PA1":  ("ADC1_IN1 (Phase 2)",   "ADC channel 1"),
    "PA6":  ("ADC1_IN6 (Phase 2)",   "ADC channel 6"),
    "PA7":  ("ADC1_IN7 (Phase 2)",   "ADC channel 7"),
    "PC0":  ("ADC1_IN10 (Phase 2)",  "ADC channel 10"),
    "PC1":  ("ADC1_IN11 (Phase 2)",  "ADC channel 11 (補位，原 PA4 讓給 DAC)"),
    # DAC1 — 專用 PA4 (G071 硬體唯一可選腳)
    "PA4":  ("DAC1_OUT1 (Phase 2)",  "DAC 輸出專用，ADC_IN4 已捨棄"),
    # GPIO bank — 通用 IO 8 隻
    "PB0":  ("GPIO bank (Phase 2)",  "通用 GPIO (亦可作 ADC_IN8)"),
    "PB1":  ("GPIO bank (Phase 2)",  "通用 GPIO (亦可作 ADC_IN9)"),
    "PB2":  ("GPIO bank (Phase 2)",  "通用 GPIO"),
    "PB10": ("GPIO bank (Phase 2)",  "通用 GPIO"),
    "PB11": ("GPIO bank (Phase 2)",  "通用 GPIO"),
    "PC6":  ("GPIO bank (Phase 2)",  "通用 GPIO"),
    "PC7":  ("GPIO bank (Phase 2)",  "通用 GPIO"),
    "PC8":  ("GPIO bank (Phase 2)",  "通用 GPIO"),
}

# 系統保留（不可動）
SYSTEM_RESERVED = {
    "PA13 (SWDIO)": "SWD 除錯資料",
    "PA14 (SWCLK)": "SWD 除錯時脈",
    "NRST":         "MCU 重置",
    "BOOT0":        "啟動模式選擇",
    "VBAT":         "電源",
    "VDDA/VREF+":   "類比電源 / 參考電壓",
    "VSSA/VREF-":   "類比接地",
    "VDD":          "數位電源",
    "VSS":          "數位接地",
    "PC14-OSC32_IN":  "LSE / GPIO (Phase 1 未用)",
    "PC15-OSC32_OUT": "LSE / GPIO (Phase 1 未用)",
    "PF0-OSC_IN":     "HSE / GPIO (Phase 1 未用)",
    "PF1-OSC_OUT":    "HSE / GPIO (Phase 1 未用)",
}


def color_for_status(status):
    """Map status to fill color (RGB)."""
    colors = {
        "Phase 1": "B7E1CD",        # 綠：active
        "Phase 2 RESERVED": "FFF2CC",  # 黃：保留位
        "SWD/Reset/Boot": "F4CCCC",  # 紅：系統關鍵
        "Power/Ground":   "D9D9D9",  # 灰：電源
        "OSC (Free)":     "DEEBF7",  # 淡藍：時脈/可用
        "Free":           "FFFFFF",  # 白：空著
    }
    return colors.get(status, "FFFFFF")


def main():
    sanity_check()

    wb = Workbook()
    ws = wb.active
    ws.title = "PinTable"

    headers = ["Pin#", "Pin Name", "Signal", "Peripheral", "Label", "Status", "備註"]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    # Data rows
    for pin_num, pin_name in LQFP64_PINS:
        # Strip alt-mode notation for matching
        bare = pin_name.split(" ")[0].split("-")[0]

        if bare in PHASE1:
            sig, peri, label, note = PHASE1[bare]
            status = "Phase 1"
        elif bare in PHASE2_RESERVED:
            sig, note = PHASE2_RESERVED[bare]
            peri = sig.split(" ")[0].split("_")[0]
            label = ""
            status = "Phase 2 RESERVED"
        elif pin_name in SYSTEM_RESERVED:
            sig = pin_name
            peri = "SYS"
            label = ""
            note = SYSTEM_RESERVED[pin_name]
            if pin_name in ("VBAT", "VDDA/VREF+", "VSSA/VREF-", "VDD", "VSS"):
                status = "Power/Ground"
            elif pin_name in ("NRST", "BOOT0", "PA13 (SWDIO)", "PA14 (SWCLK)"):
                status = "SWD/Reset/Boot"
            else:
                status = "OSC (Free)"
        else:
            sig = ""
            peri = ""
            label = ""
            note = ""
            status = "Free"

        row = [pin_num, pin_name, sig, peri, label, status, note]
        ws.append(row)

        # Apply row coloring
        fill = PatternFill("solid", fgColor=color_for_status(status))
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=ws.max_row, column=col)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths
    widths = [6, 22, 28, 12, 16, 20, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["項目", "內容"])
    ws2.append(["專案", "CommBench (G071 NUCLEO 萬用通訊測試板)"])
    ws2.append(["MCU", "STM32G071RBT6"])
    ws2.append(["封裝", "LQFP64"])
    ws2.append(["板子", "NUCLEO-G071RB"])
    ws2.append(["階段", "Phase 1 — I2C only"])
    ws2.append(["", ""])
    ws2.append(["Phase 1 啟用", "LPUART1 (VCP), I2C1 (DUT), LD4, B1"])
    ws2.append(["Phase 2 預留", "USART1 PC4/PC5, SPI1 PB3-5 + CS PB12-15, ADC1 6ch (PA0/1/6/7+PC0/1), DAC1 OUT1 PA4 專用, GPIO bank 8隻"])
    ws2.append(["ADC vs SPI", "完全分開：SPI 走 PB3-5 alt pins，ADC 全在 PA bank + PC0/PC1"])
    ws2.append(["ADC vs DAC", "PA4 永久給 DAC，ADC_IN4 已捨棄，第 6 ch 改用 PC1 (ADC_IN11)"])
    ws2.append(["", ""])
    ws2.append(["命令介面", "LPUART1 @ 115200 8N1, SCPI 風格"])
    ws2.append(["I2C 速度", "100 kHz (Timing=0x00503D58 @ 16MHz)"])
    ws2.append(["", ""])
    ws2.append(["待手動處理", "外部 I2C pull-up 4.7k 到 3V3 (PB8/PB9)"])
    ws2.append(["", ""])
    ws2.append(["Datasheet", "STM32G071RB DS12232 Rev 6"])
    ws2.append(["Board UM", "NUCLEO-G071RB UM2324"])

    for col in range(1, 3):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 70
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
        for c in row:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border

    out = r"C:\Users\jackhung\Desktop\AI\jack-toolkit\tools\CommBench\G071_NUCELO_CommBench\CommBench_PinTable.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
